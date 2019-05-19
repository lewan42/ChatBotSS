import openpyxl

class XLSXExporter():

    def getSchedule(self, path, day, week):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.worksheets[0]
        start = 4 + 12 * (day - 1)
        end = 15 + 12 * (day - 1)
        text = ""
        check_array = XLSXExporter().searchMergeCells(path)
        if week == 1:
            for row in range(start, end):
                for column in range(1, sheet.max_column):
                    value = sheet.cell(row, column).value
                    if value:
                        if row == start and column == 1:
                            text += "Н  е  д  е  л  я  №  " + str(week) + "\n\n"
                            text += str(value) + "\n"
                        if row % 2 == 0 and column != 1:
                            text += "\n"
                            text += str(value)
        if week == 2:
            for row in range(start, end):
                for column in range(1, sheet.max_column):
                    value = sheet.cell(row, column).value
                    if value:
                        if row == start and column == 1:
                            text += "Н  е  д  е  л  я  №  " + str(week) + "\n\n"
                            text += str(value) + "\n"
                        if column == 2:
                            text += "\n"
                            text += str(value)
                        if str(row) in check_array and column == 3:
                            text += "\n"
                            text += str(value)
                        if row % 2 != 0 and column != 1:
                            text += "\n"
                            text += str(value)
        return text

    def searchMergeCells(self, path):
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.worksheets[0]
        arr = worksheet.merged_cell_ranges
        check_array = []
        for item in arr:
            txt = str(item).split(":")
            first1 = txt[0][0]
            first2 = ""
            for i in range(1, len(txt[0])):
                first2 += txt[0][i]
            second1 = txt[1][0]
            second2 = ""
            for i in range(1, len(txt[1])):
                second2 += txt[1][i]
            if first1 != second1 and first2 != second2:
                check_array.append(first2)

        return check_array





